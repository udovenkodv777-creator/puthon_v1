import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_mining_model():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расчет потребности ЭКГ-10"

    # --- СТИЛИ ---
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- ШАПКА ТАБЛИЦЫ ---
    headers = ["Параметр", "Ед. изм.", "БелАЗ-75313 (130т)", "NHL TR100 (90т)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # --- ИСХОДНЫЕ ДАННЫЕ (Константы и вводные) ---
    raw_data = [
        ["Вместимость ковша ЭКГ-10", "м3", 10, 10],
        ["Коэф. наполнения ковша", "ед.", 0.85, 0.85],
        ["Коэф. разрыхления породы", "ед.", 1.3, 1.3],
        ["Плотность породы (в целике)", "т/м3", 2.6, 2.6],
        ["Грузоподъемность самосвала", "т", 130, 90],
        ["Время одного цикла экскаватора", "сек", 30, 28],
        ["Время на смену самосвала в забое", "сек", 50, 40],
        ["Годовой фонд времени работы", "час", 5500, 5500],
        ["Плановый грузооборот (2029 г.)", "тыс. тн*км", 256000, 256000],
        ["Среднее плечо транспортировки", "км", 4.5, 4.5]
    ]

    for row in raw_data:
        ws.append(row)

    # --- БЛОК РАСЧЕТА (ФОРМУЛЫ) ---
    calc_start_row = len(raw_data) + 3
    ws.cell(row=calc_start_row, column=1, value="РЕЗУЛЬТАТЫ РАСЧЕТА").font = Font(bold=True, color="FF0000")

    results_rows = [
        ("Объем горной массы в ковше", "т", "=(C2*C3*C5)/C4"),
        ("Необходимое кол-во ковшей", "шт", "=ROUNDUP(C6/C12, 0)"),
        ("Время погрузки 1 самосвала", "сек", "=(C13*C7)+C8"),
        ("Техн. производительность ЭКГ", "т/час", "=(3600/C14)*C6"),
        ("Годовая производительность ЭКГ", "тыс. т/год", "=(C15*C9)/1000"),
        ("Требуемый объем добычи (масса)", "тыс. т", "=C10/C11"),
        ("ПОТРЕБНОСТЬ В ЭКСКАВАТОРАХ", "ед.", "=ROUNDUP(C17/C16, 1)")
    ]

    curr_row = calc_start_row + 1
    for label, unit, formula in results_rows:
        ws.cell(row=curr_row, column=1, value=label)
        ws.cell(row=curr_row, column=2, value=unit)

        # Формулы для БелАЗ (колонка C) и NHL (колонка D)
        # Автоматически заменяем 'C' на 'D' для второй колонки
        ws.cell(row=curr_row, column=3, value=formula)
        ws.cell(row=curr_row, column=4, value=formula.replace("C", "D"))

        # Красим ячейки результата
        ws.cell(row=curr_row, column=3).fill = calc_fill
        ws.cell(row=curr_row, column=4).fill = calc_fill
        curr_row += 1

    # --- ОФОРМЛЕНИЕ ---
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    for row in ws.iter_rows(min_row=1, max_row=curr_row - 1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            if cell.column > 1:
                cell.alignment = center_align

    # Сохранение
    file_name = "Calculation_EKG_Model.xlsx"
    wb.save(file_name)
    print(f"Файл '{file_name}' успешно создан!")


if __name__ == "__main__":
    create_mining_model()