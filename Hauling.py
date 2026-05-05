import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Создаем рабочую книгу
wb = openpyxl.Workbook()
wb.remove(wb.active)  # удаляем дефолтный лист

# ========== Лист 1: Вводные данные ==========
ws1 = wb.create_sheet("Вводные данные")

data_input = [
    ["Параметр", "Обозначение", "Значение", "Ед. изм.", "Примечание"],
    ["Грузоподъемность БелАЗ-7513", "q_Б", 130, "т", ""],
    ["Грузоподъемность NHL TR100", "q_N", 90, "т", ""],
    ["Текущий парк БелАЗ-7513", "N_тек", 20, "ед.", ""],
    ["Плановый грузооборот (база)", "Q_баз", 140, "млн т·км", ""],
    ["Коэффициент использования парка", "k_исп", 0.85, "", ""],
    ["Годовой пробег 1 самосвала", "L_год", 120000, "км", ""],
    ["Среднее плечо транспортировки", "L_ср", 4.5, "км", ""],
    ["Время рейса", "t_рейс", 1.5, "час", ""],
    ["Производительность ЭКГ-10", "P_экс", 5.0, "млн м³/год", ""],
    ["Плотность ГМ", "ρ", 2.6, "т/м³", ""],
    ["Часов в году", "T_год", 8760, "час", ""],
]

for row in data_input:
    ws1.append(row)

# Стили для заголовков
for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# ========== Лист 2: Расчет самосвалов ==========
ws2 = wb.create_sheet("Расчет самосвалов")

# Заголовки
ws2.append(["Показатель", "Формула", "БелАЗ-7513", "NHL TR100", "Ед. изм."])

# Забираем значения из первого листа
q_B = ws1.cell(row=2, column=3).value
q_N = ws1.cell(row=3, column=3).value
k_isp = ws1.cell(row=6, column=3).value
L_god = ws1.cell(row=7, column=3).value
L_sr = ws1.cell(row=8, column=3).value
t_reys = ws1.cell(row=9, column=3).value
T_god = ws1.cell(row=12, column=3).value
N_tek = ws1.cell(row=4, column=3).value

q_reys_B = q_B * k_isp
q_reys_N = q_N * k_isp
reys_god = T_god * k_isp / t_reys
q_B_god_tkm = q_reys_B * L_god / 1000  # тыс т·км
q_N_god_tkm = q_reys_N * L_god / 1000
Q_baz = ws1.cell(row=5, column=3).value * 1000000  # в т·км

N_potr_B = Q_baz / (q_reys_B * L_god * k_isp)
N_potr_N = Q_baz / (q_reys_N * L_god * k_isp)

# Упрощенная формула
N_NHL_po_Б = N_tek * (q_B / q_N)

ws2.append(["Рейсовая нагрузка", f"=грузоподъемность × {k_isp}", q_reys_B, q_reys_N, "т"])
ws2.append(["Рейсов в год на 1 самосвал", f"=T_год × k_исп / t_рейс", int(reys_god), int(reys_god), "рейсов/год"])
ws2.append(["Грузооборот 1 самосвала в год", "=рейсовая нагрузка × L_год", q_B_god_tkm, q_N_god_tkm, "тыс. т·км"])
ws2.append(
    ["Потребное кол-во на базовый объем", "=Q_баз / (q_рейс × L_год × k_исп)", round(N_potr_B, 1), round(N_potr_N, 1),
     "ед."])
ws2.append(["То же (упрощенно)", f"=N_тек × (q_Б / q_N)", N_tek, round(N_NHL_po_Б, 1), "ед."])

# ========== Лист 3: Прирост грузооборота ==========
ws3 = wb.create_sheet("Прирост грузооборота")

ws3.append(["Год", "Прирост т·км, млн", "Прирост ГМ, млн м³", "Прирост ГМ, млн т", "Доп. самосвалы NHL",
            "Доп. экскаваторы ЭКГ-10"])
rho = ws1.cell(row=11, column=3).value
P_exc = ws1.cell(row=10, column=3).value

growth_data = [
    [2027, 76, 3.1, None, None, None],
    [2029, 125, 1.8, None, None, None],
    ["2030+", 250, 6.2, None, None, None],
]

for row_data in growth_data:
    # Расчет прироста ГМ в млн т
    V_GM = row_data[2]  # млн м³
    Q_GM_mt = V_GM * rho
    row_data[3] = round(Q_GM_mt, 2)

    # Доп. самосвалы NHL (упрощенно: 1 самосвал ≈ 0,65 млн т/год)
    # Более точно: годовская производительность 1 самосвала NHL в т
    q_N_year_mt = q_reys_N * reys_god / 1000000  # млн т/год
    N_dop = Q_GM_mt / q_N_year_mt
    row_data[4] = round(N_dop, 1)

    # Доп. экскаваторы: прирост ГМ в млн м³ / производительность экскаватора
    N_exc_dop = V_GM / P_exc
    row_data[5] = round(N_exc_dop, 2)

    ws3.append(row_data)

ws3.append([])
ws3.append(["ИТОГО доп. самосвалов NHL на развитие", "=СУММ(E2:E4)", "ед."])
ws3.append(["ИТОГО доп. экскаваторов ЭКГ-10 на развитие", "=СУММ(F2:F4)", "ед."])

# ========== Лист 4: Потребность в экскаваторах (итоговая) ==========
ws4 = wb.create_sheet("Потребность в экскаваторах")

ws4.append(["Показатель", "БелАЗ-7513", "NHL TR100", "Ед. изм.", "Примечание"])
ws4.append(["Парк самосвалов (база)", 20, 28, "ед.", "на базовый объем"])
ws4.append(["Парк самосвалов (развитие)", 54, 75, "ед.", "база + доп. на развитие"])
ws4.append(["Суммарный вывоз ГМ, млн т/год", None, None, "млн т", "расчетно"])
ws4.append(["Суммарный вывоз ГМ, млн м³/год", None, None, "млн м³", "расчетно"])
ws4.append(["Потребность в экскаваторах ЭКГ-10", None, None, "ед.", "= V_ГМ / P_экс"])

# Расчет для БелАЗ
N_B_full = 54
V_B_m3 = (N_B_full * q_reys_B * reys_god / (rho * 1000000))  # млн м³
N_exc_B = V_B_m3 / P_exc

ws4.cell(row=7, column=2, value=round(V_B_m3, 1))
ws4.cell(row=8, column=2, value=round(N_exc_B, 2))

# Расчет для NHL
N_N_full = 75
V_N_m3 = (N_N_full * q_reys_N * reys_god / (rho * 1000000))
N_exc_N = V_N_m3 / P_exc

ws4.cell(row=7, column=3, value=round(V_N_m3, 1))
ws4.cell(row=8, column=3, value=round(N_exc_N, 2))

# Итоговая строка
ws4.append([])
ws4.append(
    ["Дополнительные экскаваторы при переходе на NHL", f"≈ {round(N_exc_N - N_exc_B, 2)}", "ед.", "относительно базы"])

# ========== Форматирование ==========
for ws in [ws1, ws2, ws3, ws4]:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 3, 40)

# Сохраняем файл
wb.save("Расчет_экскаваторов_NHL_TR100.xlsx")
print("Файл 'Расчет_экскаваторов_NHL_TR100.xlsx' успешно создан!")